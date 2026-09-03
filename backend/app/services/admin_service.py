"""
Admin & super-admin business logic.
"""
from typing import List, Optional
from fastapi import HTTPException, UploadFile
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.email import send_credentials_email
from app.core.security import hash_password, generate_temp_password
from app.models.department import AdminDepartmentMapping, AdminProfile, Department
from app.models.student import StudentProfile
from app.models.user import User, UserRole, UserStatus, RegistrationSource
from app.schemas.admin import (
    AddStudentRequest, AdminCreateRequest, AdminUpdateRequest,
    DepartmentCreate, DepartmentUpdate,
    ImportPreviewResponse, ImportRowResult, ImportConfirmResponse,
)


# ── Departments ───────────────────────────────────────────────────────────────

def list_departments(db: Session) -> list:
    depts = db.query(Department).order_by(Department.department_name).all()
    result = []
    for d in depts:
        student_count = db.query(func.count(StudentProfile.id)).filter(
            StudentProfile.department_id == d.id
        ).scalar()
        admin_count = db.query(func.count(AdminDepartmentMapping.id)).filter(
            AdminDepartmentMapping.department_id == d.id
        ).scalar()
        result.append({**d.__dict__, "student_count": student_count, "admin_count": admin_count})
    return result


def create_department(db: Session, data: DepartmentCreate, created_by: str) -> Department:
    if db.query(Department).filter(Department.department_code == data.department_code).first():
        raise HTTPException(status_code=400, detail="Department code already exists")
    dept = Department(**data.model_dump(), created_by=created_by)
    db.add(dept)
    db.commit()
    db.refresh(dept)
    return dept


def update_department(db: Session, dept_id: str, data: DepartmentUpdate) -> Department:
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(dept, field, value)
    db.commit()
    db.refresh(dept)
    return dept


def delete_department(db: Session, dept_id: str) -> dict:
    dept = db.query(Department).filter(Department.id == dept_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")
    student_count = db.query(func.count(StudentProfile.id)).filter(
        StudentProfile.department_id == dept_id
    ).scalar()
    if student_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete department with {student_count} students"
        )
    db.delete(dept)
    db.commit()
    return {"message": "Department deleted"}


# ── Admin accounts ────────────────────────────────────────────────────────────

def list_admins(db: Session) -> list:
    users = db.query(User).filter(
        User.role.in_([UserRole.DEPT_ADMIN, UserRole.SUPER_ADMIN])
    ).order_by(User.created_at.desc()).all()

    result = []
    for u in users:
        profile = u.admin_profile
        if not profile:
            continue
        depts = [
            m.department for m in db.query(AdminDepartmentMapping).filter(
                AdminDepartmentMapping.admin_user_id == u.id
            ).all()
        ]
        result.append({
            "id": u.id,
            "email": u.email,
            "role": u.role.value,
            "status": u.status.value,
            "full_name": profile.full_name,
            "phone_number": profile.phone_number,
            "departments": depts,
            "last_login_at": u.last_login_at,
            "created_at": u.created_at,
        })
    return result


def create_admin(db: Session, data: AdminCreateRequest, created_by: str) -> dict:
    if db.query(User).filter(User.email == data.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")

    temp_password = generate_temp_password()

    user = User(
        email=data.email,
        password_hash=hash_password(temp_password),
        role=UserRole(data.role),
        status=UserStatus.ACTIVE,
        email_verified=True,
        must_change_password=True,
        registration_source=RegistrationSource.ADMIN_ADDED,
        created_by=created_by,
    )
    db.add(user)
    db.flush()

    profile = AdminProfile(
        user_id=user.id,
        full_name=data.full_name,
        phone_number=data.phone_number,
    )
    db.add(profile)
    db.flush()

    for dept_id in data.department_ids:
        db.add(AdminDepartmentMapping(
            admin_user_id=user.id,
            department_id=dept_id,
            assigned_by=created_by,
        ))

    db.commit()
    send_credentials_email(data.email, data.full_name, temp_password)
    return {"id": user.id, "message": "Admin account created. Credentials emailed."}


def update_admin(db: Session, admin_user_id: str, data: AdminUpdateRequest) -> dict:
    user = db.query(User).filter(User.id == admin_user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Admin not found")

    profile = user.admin_profile
    if data.full_name and profile:
        profile.full_name = data.full_name
    if data.phone_number and profile:
        profile.phone_number = data.phone_number
    if data.status:
        user.status = UserStatus(data.status)
    if data.department_ids is not None:
        db.query(AdminDepartmentMapping).filter(
            AdminDepartmentMapping.admin_user_id == admin_user_id
        ).delete()
        for dept_id in data.department_ids:
            db.add(AdminDepartmentMapping(admin_user_id=admin_user_id, department_id=dept_id))

    db.commit()
    return {"message": "Admin updated"}


def reset_admin_password(db: Session, admin_user_id: str) -> dict:
    user = db.query(User).filter(User.id == admin_user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Admin not found")
    temp_password = generate_temp_password()
    user.password_hash = hash_password(temp_password)
    user.must_change_password = True
    db.commit()
    send_credentials_email(user.email, user.admin_profile.full_name if user.admin_profile else "", temp_password)
    return {"message": "Password reset. New credentials emailed."}


# ── Student management ────────────────────────────────────────────────────────

def list_students(
    db: Session,
    department_id: Optional[str] = None,
    year: Optional[int] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    query = (
        db.query(StudentProfile, User, Department)
        .join(User, User.id == StudentProfile.user_id)
        .join(Department, Department.id == StudentProfile.department_id)
    )
    if department_id:
        query = query.filter(StudentProfile.department_id == department_id)
    if year:
        query = query.filter(StudentProfile.current_year == year)
    if search:
        term = f"%{search.lower()}%"
        query = query.filter(
            func.lower(StudentProfile.full_name).like(term) |
            func.lower(StudentProfile.roll_number).like(term) |
            func.lower(User.email).like(term)
        )

    total = query.count()
    rows = query.offset((page - 1) * page_size).limit(page_size).all()

    items = [
        {
            "user_id": user.id,
            "profile_id": profile.id,
            "full_name": profile.full_name,
            "roll_number": profile.roll_number,
            "email": user.email,
            "department_name": dept.department_name,
            "department_code": dept.department_code,
            "current_year": profile.current_year,
            "current_semester": profile.current_semester,
            "profile_completion_percentage": float(profile.profile_completion_percentage),
            "status": user.status.value,
        }
        for profile, user, dept in rows
    ]
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": max(1, -(-total // page_size)),
    }


def add_student(db: Session, data: AddStudentRequest, created_by: str) -> dict:
    if db.query(User).filter(User.email == data.email).first():
        raise HTTPException(status_code=400, detail="Email already registered")
    if db.query(StudentProfile).filter(StudentProfile.roll_number == data.roll_number).first():
        raise HTTPException(status_code=400, detail="Roll number already exists")

    temp_password = generate_temp_password()
    user = User(
        email=data.email,
        password_hash=hash_password(temp_password),
        role=UserRole.STUDENT,
        status=UserStatus.ACTIVE,
        email_verified=True,
        must_change_password=True,
        registration_source=RegistrationSource.ADMIN_ADDED,
        created_by=created_by,
    )
    db.add(user)
    db.flush()

    profile = StudentProfile(
        user_id=user.id,
        full_name=data.full_name,
        roll_number=data.roll_number,
        department_id=data.department_id,
        current_year=data.current_year,
        current_semester=data.current_semester,
        phone_number=data.phone_number,
    )
    db.add(profile)
    db.commit()

    send_credentials_email(data.email, data.full_name, temp_password)
    return {"user_id": user.id, "message": "Student account created. Credentials emailed."}


# ── Excel import ──────────────────────────────────────────────────────────────

def preview_excel_import(db: Session, file_content: bytes, filename: str) -> ImportPreviewResponse:
    """Parse xlsx/csv and validate rows without writing to DB."""
    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")

    # Expect header: full_name, email, roll_number, phone, department_code, year, semester
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    required = {"full_name", "email", "roll_number", "department_code", "year"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}"
        )

    def col(row, name):
        try:
            idx = headers.index(name)
            return str(row[idx]).strip() if row[idx] is not None else ""
        except (ValueError, IndexError):
            return ""

    results = []
    seen_emails = set()
    seen_rolls = set()

    for i, row in enumerate(rows[1:], start=2):
        email = col(row, "email")
        roll = col(row, "roll_number")
        name = col(row, "full_name")
        dept_code = col(row, "department_code")

        errors = []
        if not email:
            errors.append("Missing email")
        elif email in seen_emails:
            errors.append("Duplicate email in file")
        elif db.query(User).filter(User.email == email).first():
            errors.append("Email already registered")

        if not roll:
            errors.append("Missing roll number")
        elif roll in seen_rolls:
            errors.append("Duplicate roll number in file")
        elif db.query(StudentProfile).filter(StudentProfile.roll_number == roll).first():
            errors.append("Roll number already exists")

        if not name:
            errors.append("Missing full name")

        dept = db.query(Department).filter(Department.department_code == dept_code).first()
        if not dept_code:
            errors.append("Missing department code")
        elif not dept:
            errors.append(f"Department '{dept_code}' not found")

        seen_emails.add(email)
        seen_rolls.add(roll)

        results.append(ImportRowResult(
            row_number=i,
            email=email or None,
            roll_number=roll or None,
            status="valid" if not errors else "error",
            error_message="; ".join(errors) if errors else None,
        ))

    valid = sum(1 for r in results if r.status == "valid")
    return ImportPreviewResponse(
        total_rows=len(results),
        valid_rows=valid,
        invalid_rows=len(results) - valid,
        preview=results,
    )


def confirm_excel_import(
    db: Session, file_content: bytes, filename: str, created_by: str
) -> ImportConfirmResponse:
    preview = preview_excel_import(db, file_content, filename)
    import openpyxl, io

    wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]

    def col(row, name):
        try:
            idx = headers.index(name)
            return str(row[idx]).strip() if row[idx] is not None else ""
        except (ValueError, IndexError):
            return ""

    valid_rows = [r for r in preview.preview if r.status == "valid"]
    imported = 0
    failed = 0

    for item in valid_rows:
        row = rows[item.row_number - 1]  # row_number is 1-based including header
        try:
            dept = db.query(Department).filter(
                Department.department_code == col(row, "department_code")
            ).first()
            temp_password = generate_temp_password()
            user = User(
                email=item.email,
                password_hash=hash_password(temp_password),
                role=UserRole.STUDENT,
                status=UserStatus.ACTIVE,
                email_verified=True,
                must_change_password=True,
                registration_source=RegistrationSource.EXCEL_IMPORT,
                created_by=created_by,
            )
            db.add(user)
            db.flush()
            profile = StudentProfile(
                user_id=user.id,
                full_name=col(row, "full_name"),
                roll_number=item.roll_number,
                department_id=dept.id,
                current_year=int(col(row, "year") or 1),
                current_semester=int(col(row, "semester") or 1) if "semester" in headers else 1,
                phone_number=col(row, "phone") if "phone" in headers else None,
            )
            db.add(profile)
            db.flush()
            send_credentials_email(item.email, col(row, "full_name"), temp_password)
            imported += 1
        except Exception:
            db.rollback()
            failed += 1

    db.commit()
    return ImportConfirmResponse(
        imported_count=imported,
        failed_count=failed,
        message=f"{imported} students imported successfully.",
    )
